import argparse
import csv
import re
import warnings

from openpyxl import load_workbook


program_to_course_map = {
    "Architecture": "ARCHT-INTRN",
    "Graduate Architecture": "MARCH-INTRN",
    "Graphic Design": "GRAPH-INTRN",
    "Industrial Design": "INDUS-INTRN",
    "Interaction Design": "IXDSN-INTRN",
    "Interior Design": "INTER-INTRN",
}
programs_with_internship = list(program_to_course_map.keys())


def row_to_dict(header, row):
    return dict(zip(header, row))


def meets_program_criteria(student):
    """Check if student meets the criteria to be added to their major's
    internship course. Criteria rely on the Latest Class Standing field, which
    is a string that goes from "First Year" up to "Fifth Year". It is not a
    _level_ per se, as in an undergrad can be a "Fifth Year" and grad students
    start in their "First Year". "First Year" != Freshman.

    Args:
        student (dict): dict of student information

    Returns:
        bool: true if ready, false otherwise
    """
    major = student["Primary Program of Study"]
    # @TODO meet with programs and confirm that this is what they want
    # INDUS wants students to finish Prof Practice, we do not repload them
    if major in (
        "Architecture",
        "Interior Design",
        "Graphic Design",
        "Interaction Design",
    ) and student["Latest Class Standing"] not in ("First Year", "Second Year"):
        return True
    if (
        major == "Graduate Architecture"
        and student["Latest Class Standing"] != "First Year"
    ):
        return True
    # default to false
    return False


def make_enrollment(student, program=None):
    """create an enrollment row if student meets general criteria
    if program is present, only returns rows for students in that program

    Args:
        student (dict): dict of student information
        program (str|None): program filter or None for all programs

    Returns:
        list|False: returns a list ready to be added to a Moodle enrollment CSV
        if the student is ready, otherwise the boolean False
    """
    # students must be actively enrolled in a program with a required internship
    major = student["Primary Program of Study"]
    if (
        major in programs_with_internship
        and student["Primary Program of Study Record Status"] == "In Progress"
        and student["CCA Email"]
    ):
        if program and major != program:
            return False
        if not meets_program_criteria(student):
            return False
        # return enrollment row
        username = re.sub("@cca.edu", "", student["CCA Email"])
        course = program_to_course_map[major]
        group = "International" if student["Is International Student"] == "Yes" else ""
        return [username, course, group]
    return False


def wd_report_to_enroll_csv(report, program):
    # silence "Workbook contains no default style" warning
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        wb = load_workbook(report)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    with open("enrollments.csv", "w") as file:
        writer = csv.writer(file)
        # write CSV header row
        writer.writerow(["username", "course1", "group1"])
        for row in rows:
            student = row_to_dict(header, row)
            e = make_enrollment(student, program)
            if e:
                writer.writerow(e)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        prog="report_to_enrol",
        description="Convert Workday internships report into Moodle enrollment CSV.",
    )
    parser.add_argument(
        "-p",
        "--program",
        required=False,
        choices=programs_with_internship,
        help="Generate enrollments for only a specific program",
    )
    parser.add_argument(
        "-r",
        "--report",
        required=False,
        default="Students_for_Internship_Review.xlsx",
        help="path to the Workday Excel file",
    )
    args = parser.parse_args()
    wd_report_to_enroll_csv(args.report, args.program)
