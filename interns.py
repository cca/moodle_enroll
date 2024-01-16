import _csv  # for typing
import argparse
import csv
import re
from typing import Any, Generator, Literal
import warnings

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


program_to_course_map: dict[str, str] = {
    "Architecture": "ARCHT-INTRN",
    "Graduate Architecture": "MARCH-INTRN",
    "Graphic Design": "GRAPH-INTRN",
    "Industrial Design": "INDUS-INTRN",
    "Interaction Design": "IXDSN-INTRN",
    "Interior Design": "INTER-INTRN",
}
programs_with_internship = list(program_to_course_map.keys())


def row_to_dict(header, row) -> dict[Any, Any]:
    return dict(zip(header, row))


def meets_program_criteria(student) -> bool:
    """Check if student meets the criteria to be added to their major's
    internship course. Criteria rely on the Latest Class Standing field, which
    is a string that goes from "First Year" up to "Fifth Year". It is not a
    _level_ per se, as in an undergrad can be a "Fifth Year" and grad students
    start in their "First Year". "First Year" != Freshman.

    Args:
        student (dict): dict of student information

    Returns:
        bool: true if ready, false otherwise
    """
    major: str = student["Primary Program of Study"]
    level: str = student["Latest Class Standing"]
    # INDUS wants students to finish Prof Practice, we do not preload them
    if (
        major
        in (
            "Architecture",
            "Interior Design",
            "Graphic Design",
            "Interaction Design",
        )
        and level == "Third Year"
    ):
        return True
    if major == "Graduate Architecture" and level != "First Year":
        return True
    # default to false
    return False


def make_enrollments(student, semester, program=None, listmode=False) -> list[Any]:
    """return enrollment rows if student meets general criteria
    if program is present, only returns rows for students in that program

    Args:
        student (dict): dict of student information
        program (str|None): program filter or None for all programs

    Returns:
        list: returns a list ready to be added to a Moodle enrollment CSV
        if the student is ready, otherwise the boolean False
    """
    # students must be actively enrolled in a program with a required internship
    major: str = student["Primary Program of Study"]
    if (
        major in programs_with_internship
        and student["Primary Program of Study Record Status"] == "In Progress"
        and student["CCA Email"]
    ):
        if program and major != program:
            return []
        if not meets_program_criteria(student):
            return []
        # return enrollment rows
        username: str = re.sub("@cca.edu", "", student["CCA Email"])
        course: str = program_to_course_map[major]
        is_intl: Literal["International", False] = (
            "International" if student["Is International Student"] == "Yes" else False
        )
        if listmode:
            return [student["Student"], student["CCA Email"]]
        if is_intl:
            # intl students need 2 enrollments so they can be in 2 groups (semester and intl)
            return [(username, course, semester), (username, course, is_intl)]
        return [(username, course, semester)]
    return []


def wd_report_to_enroll_csv(args) -> None:
    # silence "Workbook contains no default style" warning
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        wb: Workbook = load_workbook(args.report)
    sheet: Worksheet = wb.worksheets[0]
    rows: Generator = sheet.iter_rows(values_only=True)
    header: tuple = next(rows)
    with open("enrollments.csv", "w") as file:
        writer: _csv._writer = csv.writer(file)
        # write CSV header row
        writer.writerow(["username", "course1", "group1"])
        for row in rows:
            student: dict[str, str] = row_to_dict(header, row)
            enrollments: list = make_enrollments(
                student, args.semester, args.program, args.list
            )
            if args.list and len(enrollments):
                print("\t".join(enrollments))
            else:
                writer.writerows(enrollments)


def semester(str) -> str:
    """validate semester string"""
    if re.match(r"(Spring|Fall|Summer) \d{4}", str):
        return str
    raise argparse.ArgumentTypeError(
        f"Semester must be in the format of 'Season YYYY' like 'Fall 2023', not '{str}'"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        prog="python interns.py",
        description="Convert Workday internships report into Moodle enrollment CSV.",
    )
    parser.add_argument(
        "-p",
        "--program",
        required=False,
        choices=programs_with_internship,
        help="Generate enrollments for only a specific program",
    )
    parser.add_argument(
        "-r",
        "--report",
        required=True,
        default="Students_for_Internship_Review.xlsx",
        help="path to the Workday Excel file",
    )
    parser.add_argument(
        "-s",
        "--semester",
        default="Fall 2024",
        type=semester,
        help='semester group (like "Fall 2023"))',
    )
    parser.add_argument(
        "-l",
        "--list",
        action="store_true",
        help="print list of students (instead of CSV)",
    )
    args = parser.parse_args()
    wd_report_to_enroll_csv(args)
